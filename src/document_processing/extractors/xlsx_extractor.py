# cockatoo_v1/src/document_processing/extractors/xlsx_extractor.py

"""
Excel XLSX file extractor using openpyxl and pandas.
"""

import os
import logging
from typing import Dict, Any, List, Union, Optional
from pathlib import Path

from .base_extractor import BaseExtractor

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not installed. Excel file processing will be limited.")

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    logger.warning("pandas not installed. Advanced Excel processing will be limited.")


class XLSXExtractor(BaseExtractor):
    """
    Excel XLSX file extractor with support for multiple sheets and formulas.
    """
    
    def __init__(self, read_formulas: bool = False, include_hidden: bool = False):
        """
        Initialize Excel extractor.
        
        Args:
            read_formulas: Whether to read cell formulas
            include_hidden: Whether to include hidden sheets/rows/columns
        """
        super().__init__()
        self.read_formulas = read_formulas
        self.include_hidden = include_hidden
        
        if not HAS_OPENPYXL:
            self.logger.warning(
                "openpyxl is not installed. Install with: pip install openpyxl"
            )
        
        if not HAS_PANDAS:
            self.logger.warning(
                "pandas is not installed. Advanced features will be limited. "
                "Install with: pip install pandas"
            )
    
    def get_supported_formats(self) -> List[str]:
        """
        Return list of supported file formats.
        
        Returns:
            List of supported file extensions
        """
        return ['.xlsx', '.xlsm', '.xltx', '.xltm', '.xls', '.xlt']
    
    def extract(self, file_path: Union[str, Path]) -> Dict[str, Any]:
        """
        Extract data from Excel file.
        
        Args:
            file_path: Path to Excel file
        
        Returns:
            Dictionary containing:
                - text: Formatted data as text
                - metadata: File metadata
                - sheets: List of sheets with data
                - formulas: Cell formulas if enabled
                - charts: Chart information
                - pivot_tables: Pivot table information
                - named_ranges: Named ranges
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        logger.info(f"Extracting Excel: {file_path}")
        
        result = {
            "text": "",
            "metadata": self.get_basic_metadata(file_path),
            "sheets": [],
            "formulas": [],
            "charts": [],
            "pivot_tables": [],
            "named_ranges": [],
            "workbook_info": {},
            "validation": {
                "is_valid": False,
                "errors": [],
            },
            "extraction_method": "openpyxl" if HAS_OPENPYXL else "pandas",
        }
        
        if HAS_OPENPYXL:
            try:
                result = self._extract_with_openpyxl(file_path, result)
                result["validation"]["is_valid"] = True
            except Exception as e:
                logger.error(f"openpyxl extraction failed: {e}")
                result["validation"]["errors"].append(str(e))
                
                # Fallback to pandas if available
                if HAS_PANDAS:
                    try:
                        result = self._extract_with_pandas(file_path, result)
                    except Exception as e2:
                        logger.error(f"pandas extraction also failed: {e2}")
                        result["validation"]["errors"].append(str(e2))
        elif HAS_PANDAS:
            try:
                result = self._extract_with_pandas(file_path, result)
                result["validation"]["is_valid"] = True
            except Exception as e:
                logger.error(f"pandas extraction failed: {e}")
                result["validation"]["errors"].append(str(e))
        else:
            result["validation"]["errors"].append(
                "No Excel library available. Install openpyxl or pandas."
            )
        
        # Generate text representation
        result["text"] = self._generate_text_representation(result)
        
        # Add language detection (Excel data is often multilingual)
        if result["text"]:
            result["metadata"]["language"] = self.detect_language(result["text"])
            result["metadata"]["summary"] = self._generate_excel_summary(result)
        
        return result
    
    def _extract_with_openpyxl(self, file_path: Path, result: Dict[str, Any]) -> Dict[str, Any]:
        """
        Extract using openpyxl library.
        
        Args:
            file_path: Path to Excel file
            result: Result dictionary to update
            
        Returns:
            Updated result dictionary
        """
        # Load workbook
        wb = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=not self.read_formulas,
            keep_vba=False,
            keep_links=False,
        )
        
        # Extract workbook metadata
        result["workbook_info"] = {
            "title": wb.properties.title or "",
            "creator": wb.properties.creator or "",
            "last_modified_by": wb.properties.lastModifiedBy or "",
            "created": str(wb.properties.created) if wb.properties.created else "",
            "modified": str(wb.properties.modified) if wb.properties.modified else "",
            "subject": wb.properties.subject or "",
            "description": wb.properties.description or "",
            "keywords": wb.properties.keywords or "",
            "category": wb.properties.category or "",
            "company": wb.properties.company or "",
            "application": wb.properties.application or "",
            "app_version": wb.properties.appVersion or "",
            "sheet_count": len(wb.sheetnames),
            "active_sheet": wb.active.title if wb.active else "",
        }
        
        # Extract sheets
        sheets = []
        formulas = []
        charts = []
        pivot_tables = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Skip hidden sheets if not included
            if not self.include_hidden and ws.sheet_state == "hidden":
                continue
            
            # Extract sheet data
            sheet_data, sheet_formulas, sheet_charts = self._extract_sheet_data(ws, sheet_name)
            
            sheets.append(sheet_data)
            formulas.extend(sheet_formulas)
            charts.extend(sheet_charts)
            
            # Detect pivot tables (simplified)
            for pivot_cache in ws._pivots:
                pivot_info = {
                    "sheet": sheet_name,
                    "cache_id": pivot_cache.cacheId,
                    "name": getattr(pivot_cache, "name", ""),
                }
                pivot_tables.append(pivot_info)
        
        # Extract named ranges
        named_ranges = []
        for name, range_def in wb.defined_names.definedName.items():
            named_ranges.append({
                "name": name,
                "range": range_def.attr_text,
            })
        
        result.update({
            "sheets": sheets,
            "formulas": formulas,
            "charts": charts,
            "pivot_tables": pivot_tables,
            "named_ranges": named_ranges,
        })
        
        # Close workbook
        wb.close()
        
        return result
    
    def _extract_sheet_data(self, worksheet, sheet_name: str):
        """
        Extract data from a worksheet.
        
        Args:
            worksheet: openpyxl Worksheet object
            sheet_name: Name of the sheet
            
        Returns:
            Tuple of (sheet_data, formulas, charts)
        """
        # Get sheet dimensions
        min_col = worksheet.min_column or 1
        max_col = worksheet.max_column or 1
        min_row = worksheet.min_row or 1
        max_row = worksheet.max_row or 1
        
        # Extract data as grid
        data_grid = []
        formulas = []
        
        for row in worksheet.iter_rows(min_row=min_row, max_row=max_row,
                                      min_col=min_col, max_col=max_col):
            row_data = []
            for cell in row:
                # Skip hidden cells if not included
                if not self.include_hidden and (cell.column in worksheet.column_dimensions and 
                                               worksheet.column_dimensions[cell.column].hidden):
                    continue
                
                # Get cell value
                cell_value = cell.value
                
                # Handle formulas
                if cell.data_type == 'f' and self.read_formulas:
                    formula = cell.value
                    if formula:
                        formulas.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "formula": formula,
                            "value": cell.value,
                        })
                    cell_value = cell.value  # Use calculated value
                
                # Format cell value
                if cell_value is None:
                    cell_value = ""
                elif isinstance(cell_value, (int, float)):
                    # Format numbers
                    cell_value = str(cell_value)
                elif isinstance(cell_value, (datetime.datetime, datetime.date)):
                    # Format dates
                    cell_value = cell_value.isoformat()
                
                row_data.append(str(cell_value))
            
            if row_data:
                data_grid.append(row_data)
        
        # Extract chart information
        charts = []
        for chart in worksheet._charts:
            chart_info = {
                "sheet": sheet_name,
                "type": chart.__class__.__name__,
                "title": getattr(chart.title, "text", "") if chart.title else "",
            }
            charts.append(chart_info)
        
        # Create sheet data structure
        sheet_data = {
            "name": sheet_name,
            "index": worksheet.sheet_format.baseColWidth,
            "data": data_grid,
            "dimensions": {
                "rows": max_row - min_row + 1,
                "columns": max_col - min_col + 1,
                "used_rows": len(data_grid),
                "used_columns": len(data_grid[0]) if data_grid else 0,
            },
            "properties": {
                "sheet_state": worksheet.sheet_state,
                "tab_color": str(worksheet.sheet_properties.tabColor) if worksheet.sheet_properties.tabColor else "",
                "page_setup": {
                    "orientation": worksheet.page_setup.orientation,
                    "paper_size": worksheet.page_setup.paperSize,
                },
            },
        }
        
        return sheet_data, formulas, charts
    
    def _extract_with_pandas(self, file_path: Path, result: Dict[str, Any]) -> Dict[str, Any]:
        """
        Extract using pandas library.
        
        Args:
            file_path: Path to Excel file
            result: Result dictionary to update
            
        Returns:
            Updated result dictionary
        """
        # Read Excel file with pandas
        try:
            # Get sheet names
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            sheets = []
            
            for sheet_name in sheet_names:
                # Read sheet
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    header=None,  # Read all data
                    dtype=str,    # Read as strings
                    na_filter=False,  # Don't convert empty strings to NaN
                )
                
                # Convert to list of lists
                data = df.values.tolist()
                
                # Clean data (replace NaN with empty strings)
                cleaned_data = []
                for row in data:
                    cleaned_row = [str(cell) if pd.notna(cell) else "" for cell in row]
                    cleaned_data.append(cleaned_row)
                
                sheet_data = {
                    "name": sheet_name,
                    "data": cleaned_data,
                    "dimensions": {
                        "rows": len(cleaned_data),
                        "columns": len(cleaned_data[0]) if cleaned_data else 0,
                    },
                }
                
                sheets.append(sheet_data)
            
            result["sheets"] = sheets
            result["workbook_info"] = {
                "sheet_count": len(sheet_names),
            }
            
        except Exception as e:
            logger.error(f"Failed to read Excel with pandas: {e}")
            raise
        
        return result
    
    def _generate_text_representation(self, result: Dict[str, Any]) -> str:
        """
        Generate text representation of Excel data.
        
        Args:
            result: Extraction result
            
        Returns:
            Text representation
        """
        text_parts = []
        
        for sheet in result.get("sheets", []):
            sheet_name = sheet.get("name", "Unknown")
            data = sheet.get("data", [])
            
            if data:
                text_parts.append(f"=== Sheet: {sheet_name} ===")
                
                # Limit rows for text representation
                max_rows_display = 50
                display_data = data[:max_rows_display]
                
                for i, row in enumerate(display_data):
                    # Join row cells with tabs
                    row_text = "\t".join(str(cell) for cell in row)
                    text_parts.append(f"Row {i+1}:\t{row_text}")
                
                if len(data) > max_rows_display:
                    text_parts.append(f"\n... and {len(data) - max_rows_display} more rows")
                
                text_parts.append("")  # Empty line between sheets
        
        return "\n".join(text_parts)
    
    def _generate_excel_summary(self, result: Dict[str, Any]) -> str:
        """
        Generate summary of Excel file.
        
        Args:
            result: Extraction result
            
        Returns:
            Summary string
        """
        sheets = result.get("sheets", [])
        
        if not sheets:
            return "Empty Excel file"
        
        summary_parts = []
        
        summary_parts.append(f"Sheets: {len(sheets)}")
        
        # Sheet information
        for sheet in sheets[:5]:  # Limit to first 5 sheets
            name = sheet.get("name", "")
            dims = sheet.get("dimensions", {})
            rows = dims.get("used_rows", 0)
            cols = dims.get("used_columns", 0)
            
            summary_parts.append(f"  - {name}: {rows} rows × {cols} columns")
        
        if len(sheets) > 5:
            summary_parts.append(f"  ... and {len(sheets) - 5} more sheets")
        
        # Total data
        total_rows = sum(s.get("dimensions", {}).get("used_rows", 0) for s in sheets)
        total_cells = sum(s.get("dimensions", {}).get("used_rows", 0) * 
                         s.get("dimensions", {}).get("used_columns", 0) for s in sheets)
        
        summary_parts.append(f"Total rows: {total_rows}")
        summary_parts.append(f"Total cells: {total_cells}")
        
        # Additional features
        if result.get("formulas"):
            summary_parts.append(f"Formulas: {len(result['formulas'])}")
        
        if result.get("charts"):
            summary_parts.append(f"Charts: {len(result['charts'])}")
        
        if result.get("pivot_tables"):
            summary_parts.append(f"Pivot tables: {len(result['pivot_tables'])}")
        
        return "; ".join(summary_parts)
    
    def extract_sheet_as_dataframe(self, file_path: Union[str, Path], 
                                  sheet_name: str = None,
                                  header_row: int = 0) -> Optional[pd.DataFrame]:
        """
        Extract sheet as pandas DataFrame.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to extract (None for first sheet)
            header_row: Row to use as header (None for no header)
            
        Returns:
            pandas DataFrame or None if failed
        """
        if not HAS_PANDAS:
            logger.error("pandas is required for DataFrame extraction")
            return None
        
        try:
            if sheet_name:
                df = pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    header=header_row,
                    engine='openpyxl'
                )
            else:
                df = pd.read_excel(
                    file_path,
                    header=header_row,
                    engine='openpyxl'
                )
            
            return df
            
        except Exception as e:
            logger.error(f"Failed to extract sheet as DataFrame: {e}")
            return None
    
    def get_sheet_statistics(self, file_path: Union[str, Path], 
                            sheet_name: str = None) -> Dict[str, Any]:
        """
        Get statistics for a sheet.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet (None for active sheet)
            
        Returns:
            Sheet statistics
        """
        result = self.extract(file_path)
        
        if not result["validation"]["is_valid"]:
            return {"error": "Invalid Excel file"}
        
        sheets = result.get("sheets", [])
        
        if sheet_name:
            target_sheet = next((s for s in sheets if s["name"] == sheet_name), None)
        else:
            target_sheet = sheets[0] if sheets else None
        
        if not target_sheet:
            return {"error": f"Sheet '{sheet_name}' not found"}
        
        data = target_sheet.get("data", [])
        dims = target_sheet.get("dimensions", {})
        
        # Calculate statistics
        stats = {
            "sheet_name": target_sheet["name"],
            "rows": dims.get("used_rows", 0),
            "columns": dims.get("used_columns", 0),
            "total_cells": dims.get("used_rows", 0) * dims.get("used_columns", 0),
            "empty_cells": 0,
            "numeric_cells": 0,
            "text_cells": 0,
            "date_cells": 0,
            "max_row_length": 0,
            "data_types": {},
        }
        
        for row in data:
            stats["max_row_length"] = max(stats["max_row_length"], len(row))
            
            for cell in row:
                cell_str = str(cell)
                
                if not cell_str.strip():
                    stats["empty_cells"] += 1
                else:
                    # Try to determine data type
                    try:
                        float(cell_str)
                        stats["numeric_cells"] += 1
                    except:
                        # Check for date
                        if any(sep in cell_str for sep in ['-', '/', ':']):
                            try:
                                pd.to_datetime(cell_str)
                                stats["date_cells"] += 1
                            except:
                                stats["text_cells"] += 1
                        else:
                            stats["text_cells"] += 1
        
        # Calculate percentages
        total_cells = stats["total_cells"]
        if total_cells > 0:
            stats["empty_percentage"] = (stats["empty_cells"] / total_cells) * 100
            stats["numeric_percentage"] = (stats["numeric_cells"] / total_cells) * 100
            stats["text_percentage"] = (stats["text_cells"] / total_cells) * 100
            stats["date_percentage"] = (stats["date_cells"] / total_cells) * 100
        
        return stats